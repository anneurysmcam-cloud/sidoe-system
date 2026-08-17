"""
data/diccionario_datos.py
==========================
Genera la hoja "Diccionario de Datos" que se agrega al Excel exportado desde
Consultas (ver views/consultas.py), documentando las hojas "Indicadores",
"Fuentes" y "Factibilidad" de ese mismo archivo.

Por qué existe este módulo
---------------------------
La ONE, a través de "Lineamientos y Recomendaciones para Documentar el
Diccionario de Datos" (Dirección de Normativas y Metodologías, v1.0,
20/08/2025), clasifica como "Diccionario de Datos Pasivo" cualquier conjunto
de datos que no vive dentro de un sistema gestor de base de datos, como un
archivo XLSX — exactamente el caso del Excel de Consultas — y exige que se
documente de forma manual (Generalidades, pág. 17 del documento). Este
módulo es esa documentación manual, generada en código para que nunca quede
desactualizada respecto a las columnas reales que se exportan.

Estructura adoptada (combina las dos referencias que compartió Randy):
  - El formato de 4 columnas (Nombre, Tipo, Etiqueta, Valores) del
    "Diccionario de Base de datos de Protección Social" (ONE, dic. 2024).
  - Los "Metadatos Técnicos" y "Metadatos Semánticos" del Anexo 2 de los
    Lineamientos (ID/Nombre de Tabla, Tipo de Dato, Condición, Dominio de
    Valores, Descripción de la Variable, Comentarios Adicionales), aplanados
    en una sola tabla por legibilidad dentro de una hoja de cálculo.

No se documentan aquí los campos personalizados de Auxiliares uno por uno
(varían por institución/configuración): se listan igual, para que el
Diccionario refleje el 100% de las columnas realmente exportadas, pero con
una descripción genérica que remite al módulo de Auxiliares como fuente de
su definición real.
"""

from __future__ import annotations

import datetime as _dt

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Catálogo de metadatos por columna, agrupado por hoja de origen.
# Cada entrada: etiqueta, tipo_dato, condicion, dominio, descripcion.
# ---------------------------------------------------------------------------

_META_INDICADORES: dict[str, dict[str, str]] = {
    "codigo": dict(
        etiqueta="Código del Indicador", tipo="Texto", condicion="Obligatorio",
        dominio="Código único alfanumérico asignado por SIDOE",
        descripcion="Identificador único del indicador dentro del sistema; "
                     "actúa como llave primaria del indicador.",
    ),
    "generador_demanda": dict(
        etiqueta="Generador de Demanda", tipo="Texto", condicion="Obligatorio",
        dominio="END, ODS, PNPSP, CMV",
        descripcion="Instrumento de política pública que origina la "
                     "necesidad de este indicador.",
    ),
    "eje": dict(
        etiqueta="Eje", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Eje temático o estratégico del generador de demanda "
                     "al que se asocia el indicador.",
    ),
    "politica_gobierno": dict(
        etiqueta="Política de Gobierno", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Política pública específica vinculada al indicador.",
    ),
    "otros_ejes_politicas": dict(
        etiqueta="Otros Ejes/Políticas", tipo="Texto", condicion="Opcional",
        dominio="Lista de valores separados por coma",
        descripcion="Ejes o políticas secundarios adicionales asociados al "
                     "mismo indicador.",
    ),
    "indicador": dict(
        etiqueta="Nombre del Indicador", tipo="Texto", condicion="Obligatorio",
        dominio="Texto libre",
        descripcion="Nombre oficial y descriptivo del indicador estadístico.",
    ),
    "indicador_referenciado": dict(
        etiqueta="Indicador Referenciado / Duplicado", tipo="Texto", condicion="Opcional",
        dominio="Código de otro indicador registrado en el sistema",
        descripcion="Señala que este indicador comparte fuente y "
                     "tratamiento metodológico con otro (mismo indicador "
                     "asociado a un generador de demanda distinto).",
    ),
    "dominio_actividad_estadistica": dict(
        etiqueta="Dominio de Actividad Estadística", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Área temática estadística del indicador (ej. "
                     "Demografía, Economía).",
    ),
    "subdominio_actividad_estadistica": dict(
        etiqueta="Subdominio de Actividad Estadística", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Subdivisión del dominio de actividad estadística.",
    ),
    "area_misional_one": dict(
        etiqueta="Área Misional ONE", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Dirección o área de la ONE responsable misionalmente "
                     "del indicador.",
    ),
    "sector_ioe": dict(
        etiqueta="Sector IOE", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Sector de la Institución/Operación Estadística (IOE) "
                     "asociado al indicador.",
    ),
    "requerimiento_clasificacion": dict(
        etiqueta="Requerimiento de Clasificación", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si el indicador requiere una clasificación "
                     "estadística estandarizada.",
    ),
    "especificar_clasificacion": dict(
        etiqueta="Especificar Clasificación", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Detalle de la clasificación estadística cuando aplica.",
    ),
    "metodo_calculo": dict(
        etiqueta="Método de Cálculo", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Descripción metodológica de cómo se calcula el "
                     "indicador.",
    ),
    "ficha_tecnica": dict(
        etiqueta="Ficha Técnica", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Contenido o referencia de la ficha técnica "
                     "metodológica del indicador.",
    ),
    "numerador": dict(
        etiqueta="Numerador", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Definición del numerador de la fórmula de cálculo.",
    ),
    "denominador": dict(
        etiqueta="Denominador", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Definición del denominador de la fórmula de cálculo.",
    ),
    "unidad_medida": dict(
        etiqueta="Unidad de Medida", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Unidad en la que se expresa el resultado del "
                     "indicador (ej. porcentaje, tasa, número absoluto).",
    ),
    "sexo_indicador": dict(
        etiqueta="Desagregación por Sexo", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si el indicador se desagrega por sexo.",
    ),
    "edad_indicador": dict(
        etiqueta="Desagregación por Edad", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si el indicador se desagrega por edad.",
    ),
    "territorio_indicador": dict(
        etiqueta="Desagregación Territorial", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si el indicador se desagrega territorialmente.",
    ),
    "discapacidad_indicador": dict(
        etiqueta="Desagregación por Discapacidad", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si el indicador se desagrega por condición de "
                     "discapacidad.",
    ),
    "nivel_ingreso_indicador": dict(
        etiqueta="Desagregación por Nivel de Ingreso", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si el indicador se desagrega por nivel de "
                     "ingreso socioeconómico.",
    ),
    "periodicidad_indicador": dict(
        etiqueta="Periodicidad del Indicador", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Frecuencia con la que se actualiza/calcula el "
                     "indicador (ej. Anual, Mensual).",
    ),
    "ente_responsable_metodologia": dict(
        etiqueta="Ente Responsable de la Metodología", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Institución responsable de definir o avalar la "
                     "metodología del indicador.",
    ),
    "alcance_metodologico": dict(
        etiqueta="Alcance Metodológico", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Cobertura conceptual y metodológica declarada para el "
                     "indicador.",
    ),
    "num_fuentes": dict(
        etiqueta="Número de Fuentes", tipo="Numérico (entero)", condicion="Calculado",
        dominio="Entero ≥ 0",
        descripcion="Cantidad de fuentes de información registradas para "
                     "este indicador; se calcula al momento de exportar.",
    ),
}

_META_FUENTES: dict[str, dict[str, str]] = {
    "indicador_codigo": dict(
        etiqueta="Código del Indicador (FK)", tipo="Texto", condicion="Obligatorio",
        dominio="Debe existir como 'codigo' en la hoja Indicadores",
        descripcion="Código del indicador al que pertenece esta fuente "
                     "(llave foránea hacia la hoja Indicadores).",
    ),
    "indicador_nombre": dict(
        etiqueta="Nombre del Indicador", tipo="Texto", condicion="Obligatorio",
        dominio="Texto libre",
        descripcion="Nombre del indicador asociado, incluido por "
                     "legibilidad de la hoja.",
    ),
    "nombre_fuente": dict(
        etiqueta="Nombre de la Fuente", tipo="Texto", condicion="Obligatorio",
        dominio="Texto libre",
        descripcion="Nombre de la fuente de datos (operación estadística o "
                     "registro administrativo) que provee el dato.",
    ),
    "tipo_fuente": dict(
        etiqueta="Tipo de Fuente", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Naturaleza de la fuente de información (ej. Encuesta, "
                     "Censo, Registro Administrativo).",
    ),
    "institucion_productora": dict(
        etiqueta="Institución Productora", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Institución responsable de producir o administrar la "
                     "fuente.",
    ),
    "periodicidad_fuente": dict(
        etiqueta="Periodicidad de la Fuente", tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Frecuencia de actualización de la fuente.",
    ),
    "existencia_fuente": dict(
        etiqueta="Existencia de la Fuente (C2.1)", tipo="Texto", condicion="Opcional",
        dominio="Completamente | Parcialmente | No hay fuente",
        descripcion="Grado en que la fuente existe y está disponible; "
                     "criterio C2.1 del Engine de Factibilidad.",
    ),
    "sexo_fuente": dict(
        etiqueta="Desagregación por Sexo (disponible en la fuente)", tipo="Texto",
        condicion="Opcional", dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si la fuente permite desagregar por sexo.",
    ),
    "edad_fuente": dict(
        etiqueta="Desagregación por Edad (disponible en la fuente)", tipo="Texto",
        condicion="Opcional", dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si la fuente permite desagregar por edad.",
    ),
    "territorio_fuente": dict(
        etiqueta="Desagregación Territorial (disponible en la fuente)", tipo="Texto",
        condicion="Opcional", dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si la fuente permite desagregar territorialmente.",
    ),
    "discapacidad_fuente": dict(
        etiqueta="Desagregación por Discapacidad (disponible en la fuente)",
        tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si la fuente permite desagregar por condición "
                     "de discapacidad.",
    ),
    "nivel_ingreso_socioeconomico": dict(
        etiqueta="Desagregación por Nivel de Ingreso (disponible en la fuente)",
        tipo="Texto", condicion="Opcional",
        dominio="Catálogo institucional (módulo Auxiliares)",
        descripcion="Indica si la fuente permite desagregar por nivel de "
                     "ingreso socioeconómico.",
    ),
    "ioe": dict(
        etiqueta="IOE (Institución/Operación Estadística)", tipo="Texto",
        condicion="Opcional", dominio="Texto libre",
        descripcion="Identificación de la Institución u Operación "
                     "Estadística de origen de la fuente.",
    ),
    "ra": dict(
        etiqueta="RA (Registro Administrativo)", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Identificación del registro administrativo de "
                     "origen, cuando la fuente es de ese tipo.",
    ),
    "calculado_datos_agregados": dict(
        etiqueta="Calculado con Datos Agregados", tipo="Texto", condicion="Opcional",
        dominio="Sí | No",
        descripcion="Indica si el indicador se calcula a partir de datos "
                     "ya agregados provistos por la fuente.",
    ),
    "hipervinculo_ultimo_calculo": dict(
        etiqueta="Hipervínculo del Último Cálculo", tipo="Texto (URL)", condicion="Opcional",
        dominio="URL válida",
        descripcion="Enlace a la publicación o cálculo más reciente "
                     "disponible de esta fuente.",
    ),
    "anio_ultimo_dato_disponible": dict(
        etiqueta="Año del Último Dato Disponible", tipo="Numérico (año)", condicion="Opcional",
        dominio="Año calendario (ej. 2025)",
        descripcion="Año más reciente con datos disponibles en esta "
                     "fuente.",
    ),
    "comentarios": dict(
        etiqueta="Comentarios", tipo="Texto", condicion="Opcional",
        dominio="Texto libre",
        descripcion="Observaciones adicionales sobre la fuente.",
    ),
}

_META_FACTIBILIDAD: dict[str, dict[str, str]] = {
    "codigo": _META_INDICADORES["codigo"],
    "indicador": _META_INDICADORES["indicador"],
    "generador_demanda": _META_INDICADORES["generador_demanda"],
    "c1_metodologia": dict(
        etiqueta="C1 — Metodología", tipo="Texto", condicion="Obligatorio",
        dominio="Vocabulario fijo del Engine (4 opciones oficiales)",
        descripcion="Criterio C1: existencia y solidez de la metodología "
                     "declarada para el indicador.",
        comentarios="Puntaje del Engine de Factibilidad — 'Indicador con "
                     "metodología nacional o internacional definida' = 15; "
                     "'…método de cálculo es auto explicativo' = 7.5; "
                     "'…método de cálculo se puede establecer mediante "
                     "criterio experto' = 7.5; 'No cumple con los criterios "
                     "anteriores' = 0. Máximo del criterio: 15.",
    ),
    "c21_existencia_fuente": dict(
        etiqueta="C2.1 — Existencia de la Fuente", tipo="Texto", condicion="Obligatorio",
        dominio="Completamente | Parcialmente | No hay fuente",
        descripcion="Criterio C2.1: grado de existencia de una fuente "
                     "para el indicador.",
        comentarios="Puntaje del Engine de Factibilidad — Completamente = "
                     "15; Parcialmente = 7.5; No hay fuente = 0. Máximo del "
                     "criterio: 15.",
    ),
    "c22_disponibilidad": dict(
        etiqueta="C2.2 — Disponibilidad", tipo="Texto", condicion="Obligatorio",
        dominio="Sí | No",
        descripcion="Criterio C2.2: disponibilidad efectiva de los datos "
                     "de la fuente.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 10; No = 0. "
                     "Máximo del criterio: 10.",
    ),
    "c23_periodicidad_establecida": dict(
        etiqueta="C2.3 — Periodicidad Establecida", tipo="Texto", condicion="Obligatorio",
        dominio="Sí | No",
        descripcion="Criterio C2.3: si la fuente tiene una periodicidad de "
                     "actualización establecida.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 10; No = 0. "
                     "Máximo del criterio: 10.",
    ),
    "c31_posee_desagregacion": dict(
        etiqueta="C3.1 — Posee Desagregación", tipo="Texto", condicion="Obligatorio",
        dominio="Sí | No | No es requerida",
        descripcion="Criterio C3.1: si la fuente permite desagregar el "
                     "indicador.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 5; No = 0; "
                     "No es requerida = 5. Máximo del criterio: 5.",
    ),
    "num_desagregaciones_requeridas": dict(
        etiqueta="Desagregaciones Requeridas (C3.2)", tipo="Numérico (entero)",
        condicion="Opcional", dominio="Entero ≥ 0",
        descripcion="Cantidad de desagregaciones que el indicador "
                     "requiere conceptualmente.",
        comentarios="No tiene puntaje por opción: junto con "
                     "'Desagregaciones Disponibles' alimenta la fórmula "
                     "C3.2 del Engine = mín(disponibles/requeridas, 1) × "
                     "5. Máximo del criterio: 5.",
    ),
    "num_desagregaciones_disponibles": dict(
        etiqueta="Desagregaciones Disponibles (C3.2)", tipo="Numérico (entero)",
        condicion="Opcional", dominio="Entero ≥ 0",
        descripcion="Cantidad de desagregaciones efectivamente "
                     "disponibles en la fuente.",
        comentarios="No tiene puntaje por opción: junto con "
                     "'Desagregaciones Requeridas' alimenta la fórmula "
                     "C3.2 del Engine = mín(disponibles/requeridas, 1) × "
                     "5. Máximo del criterio: 5.",
    ),
    "articulacion_fuentes": dict(
        etiqueta="Articulación de Fuentes", tipo="Texto", condicion="Opcional",
        dominio="Sí se articula | No se articula | No requiere de "
                 "articulación",
        descripcion="Grado de articulación entre múltiples fuentes del "
                     "indicador, cuando aplica.",
        comentarios="Puntaje del Engine de Factibilidad — Sí se articula = "
                     "6.667; No requiere de articulación = 6.667; No se "
                     "articula = 0. Máximo del criterio: 6.667.",
    ),
    "armonizacion_conceptual": dict(
        etiqueta="Armonización Conceptual", tipo="Texto", condicion="Opcional",
        dominio="Sí | No",
        descripcion="Indica si existen brechas de armonización conceptual "
                     "entre fuentes.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 0; No = "
                     "6.667 (penaliza la presencia de brechas). Máximo del "
                     "criterio: 6.667.",
    ),
    "subregistro_cobertura": dict(
        etiqueta="Subregistro de Cobertura", tipo="Texto", condicion="Opcional",
        dominio="Sí | No",
        descripcion="Indica si la fuente presenta subregistro de "
                     "cobertura.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 0; No = "
                     "6.667 (penaliza la presencia de subregistro). "
                     "Máximo del criterio: 6.667.",
    ),
    "cobertura_territorial": dict(
        etiqueta="Cobertura Territorial", tipo="Texto", condicion="Opcional",
        dominio="Sí | No",
        descripcion="Indica si la fuente tiene cobertura territorial "
                     "completa.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 6.667; No "
                     "= 0. Máximo del criterio: 6.667.",
    ),
    "estructura_datos": dict(
        etiqueta="Estructura de Datos", tipo="Texto", condicion="Opcional",
        dominio="3 opciones oficiales (base de datos estructurada / "
                 "formato Excel / ninguna)",
        descripcion="Tipo de estructura de datos que utiliza la fuente en "
                     "su procesamiento.",
        comentarios="Puntaje del Engine de Factibilidad — 'La fuente "
                     "utiliza en el procesamiento una base de datos "
                     "estructurada' = 6.667; 'No posee una base de datos "
                     "estructurada, pero posee un formato para montar "
                     "datos (Excel)' = 3.3335; ninguna de las anteriores = "
                     "0. Máximo del criterio: 6.667.",
    ),
    "variables_calculo": dict(
        etiqueta="Variables de Cálculo", tipo="Texto", condicion="Opcional",
        dominio="Sí | No | No identificada | No requerida",
        descripcion="Indica si están identificadas las variables "
                     "necesarias para el cálculo del indicador.",
        comentarios="Puntaje del Engine de Factibilidad — Sí = 6.667; No "
                     "= 0; No identificada = 6.667; No requerida = 6.667. "
                     "Máximo del criterio: 6.667.",
    ),
    "puntaje": dict(
        etiqueta="Puntaje de Factibilidad", tipo="Numérico (decimal)", condicion="Calculado",
        dominio="0 a 100",
        descripcion="Puntaje final calculado por el Engine de "
                     "Factibilidad a partir de los criterios C1 a C3.",
    ),
    "factibilidad": dict(
        etiqueta="Categoría de Factibilidad", tipo="Texto", condicion="Calculado",
        dominio="Factibilidad I (puntaje ≥ 91) | Factibilidad II "
                 "(puntaje ≥ 70) | Factibilidad III (puntaje < 70 o sin "
                 "datos)",
        descripcion="Categoría resultante del puntaje: I = alta "
                     "factibilidad, II = media, III = baja.",
    ),
    "calc_timestamp": dict(
        etiqueta="Fecha del Último Cálculo", tipo="Fecha y hora", condicion="Calculado",
        dominio="Fecha/hora local de República Dominicana",
        descripcion="Momento del último cálculo de factibilidad para el "
                     "indicador (convertido de UTC a hora local de RD).",
    ),
}

_METADATOS_POR_HOJA: dict[str, dict[str, dict[str, str]]] = {
    "Indicadores": _META_INDICADORES,
    "Fuentes": _META_FUENTES,
    "Factibilidad": _META_FACTIBILIDAD,
}

_COLUMNAS_TABLA = [
    "Hoja", "Nombre de la Variable", "Etiqueta", "Tipo de Dato",
    "Condición", "Dominio / Valores Permitidos", "Descripción",
    "Comentarios Adicionales",
]


def _fila_para_columna(hoja: str, columna: str) -> dict[str, str]:
    """Devuelve la fila del diccionario para una columna dada, con fallback
    genérico para campos personalizados de Auxiliares no documentados
    estáticamente (varían por institución/configuración)."""
    meta = _METADATOS_POR_HOJA.get(hoja, {}).get(columna)
    if meta is None:
        return dict(
            Hoja=hoja,
            **{"Nombre de la Variable": columna},
            Etiqueta=columna.replace("_", " ").strip().title(),
            **{"Tipo de Dato": "Texto"},
            Condición="Opcional",
            **{"Dominio / Valores Permitidos": "Catálogo personalizado (módulo Auxiliares)"},
            Descripción=(
                "Campo personalizado configurado por la institución a "
                "través del módulo de Auxiliares; su definición vive en "
                "ese catálogo, no en el vocabulario fijo del sistema."
            ),
            **{"Comentarios Adicionales": ""},
        )
    return dict(
        Hoja=hoja,
        **{"Nombre de la Variable": columna},
        Etiqueta=meta["etiqueta"],
        **{"Tipo de Dato": meta["tipo"]},
        Condición=meta["condicion"],
        **{"Dominio / Valores Permitidos": meta["dominio"]},
        Descripción=meta["descripcion"],
        **{"Comentarios Adicionales": meta.get("comentarios", "")},
    )


def construir_tabla_diccionario_datos(
    df_indicadores: pd.DataFrame,
    df_fuentes: pd.DataFrame,
    df_factibilidad: pd.DataFrame,
) -> pd.DataFrame:
    """Arma la tabla del Diccionario de Datos a partir de las columnas
    REALMENTE presentes en cada DataFrame exportado (incluye campos
    personalizados de Auxiliares si están activos), en el mismo orden en
    que aparecen en cada hoja del Excel."""
    filas = []
    for hoja, df in (
        ("Indicadores", df_indicadores),
        ("Fuentes", df_fuentes),
        ("Factibilidad", df_factibilidad),
    ):
        for columna in df.columns:
            filas.append(_fila_para_columna(hoja, columna))
    return pd.DataFrame(filas, columns=_COLUMNAS_TABLA)


# ---------------------------------------------------------------------------
# Escritura en el Excel: bloque de identificación institucional (formato
# "Ficha técnica" de los Lineamientos ONE) + tabla, con formato openpyxl.
# ---------------------------------------------------------------------------

_FILL_ENCABEZADO = PatternFill("solid", fgColor="0F3A7A")
_FONT_ENCABEZADO = Font(color="FFFFFF", bold=True, name="Arial")
_FONT_TITULO = Font(bold=True, size=13, name="Arial", color="0F3A7A")
_FONT_ETIQUETA_FICHA = Font(bold=True, name="Arial")
_FONT_NORMAL = Font(name="Arial")
_ANCHO_COLUMNAS = [14, 30, 34, 18, 14, 46, 60, 40]


def aplicar_formato_encabezado_hoja_datos(
    ws: Worksheet, num_columnas: int, congelar_panel: bool = True
) -> None:
    """Aplica el mismo estilo de encabezado azul/blanco institucional de
    la hoja "Diccionario de Datos" (fondo #0F3A7A, texto blanco en
    negrita) a la fila 1 de una hoja ya escrita con
    ``DataFrame.to_excel(writer, ..., sheet_name=...)`` — pensado para
    unificar la identidad visual de las hojas "Indicadores", "Fuentes" y
    "Factibilidad" del Excel exportado con la del propio Diccionario de
    Datos (pedido explícito de Randy: "tiene mucho flow").

    Se asume que la fila de encabezados quedó en la fila 1, que es lo que
    hace ``to_excel`` por defecto sin ``startrow``. Para hojas con un
    bloque de identificación antes de la tabla (como la propia
    "Diccionario de Datos"), usar el formateo manual de
    ``escribir_hoja_diccionario_datos`` en su lugar — esta función es para
    el caso simple de una hoja que ES la tabla desde la fila 1.
    """
    for col_idx in range(1, num_columnas + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.fill = _FILL_ENCABEZADO
        celda.font = _FONT_ENCABEZADO
        celda.alignment = Alignment(wrap_text=True, vertical="center")
    if congelar_panel:
        ws.freeze_panes = ws.cell(row=2, column=1)


def ajustar_ancho_columnas_auto(
    ws: Worksheet, df: pd.DataFrame, ancho_min: int = 10, ancho_max: int = 45
) -> None:
    """Ajusta el ancho de cada columna de ``ws`` al contenido real de ``df``
    (encabezado + valores), para evitar el recorte visual de texto que se ve
    con el ancho fijo por defecto de openpyxl (~8.4 para todas las
    columnas, sin importar el contenido). El ancho se acota entre
    ``ancho_min`` y ``ancho_max`` para que columnas de texto muy largo (ej.
    Método de Cálculo, Ficha Técnica) no desborden la hoja; ese contenido
    queda con ``wrap_text`` activado en las celdas de datos para seguir
    siendo legible sin una columna gigante.

    Pensado para aplicarse a las hojas "Indicadores", "Fuentes" y
    "Factibilidad" del Excel exportado desde Consultas, después de
    ``aplicar_formato_encabezado_hoja_datos``.
    """
    for idx, columna in enumerate(df.columns, start=1):
        # No usar Series.astype(str).map(len): con el backend Arrow de
        # pandas 3.x, los valores nulos (NaN/None) de columnas numéricas u
        # object no siempre quedan convertidos a texto por astype(str) —
        # algunos llegan a .map() como float, y len(float) lanza TypeError
        # (bug reportado por Randy, ERR-1786666895). Se recorre valor por
        # valor con pd.isna() como guardia explícita.
        if len(df):
            max_len_datos = max(
                (len(str(v)) for v in df[columna] if pd.notna(v)), default=0
            )
        else:
            max_len_datos = 0
        # Ancho del encabezado sin acotar (para que siempre quepa en una
        # sola línea, pedido explícito de Randy) + ancho de los datos
        # acotado entre ancho_min y ancho_max para que un valor largo no
        # desborde la hoja.
        ancho_encabezado = len(str(columna)) + 4
        ancho_datos = max(ancho_min, min(max_len_datos + 2, ancho_max))
        ancho = max(ancho_encabezado, ancho_datos)
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    ultima_fila = 1 + len(df)
    for fila in ws.iter_rows(min_row=2, max_row=ultima_fila, max_col=len(df.columns)):
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, vertical="top")


def escribir_hoja_diccionario_datos(
    writer: pd.ExcelWriter,
    df_indicadores: pd.DataFrame,
    df_fuentes: pd.DataFrame,
    df_factibilidad: pd.DataFrame,
    sheet_name: str = "Diccionario de Datos",
) -> None:
    """Escribe la hoja "Diccionario de Datos" dentro del ``writer`` de
    pandas/openpyxl ya abierto, con un encabezado de identificación
    institucional (estilo "Ficha técnica" de los Lineamientos ONE) seguido
    de la tabla de metadatos técnicos y semánticos de cada hoja exportada.
    """
    tabla = construir_tabla_diccionario_datos(
        df_indicadores, df_fuentes, df_factibilidad
    )

    FILA_INICIO_TABLA = 9  # deja espacio (0-indexed) para el bloque de ficha técnica
    tabla.to_excel(
        writer, index=False, sheet_name=sheet_name, startrow=FILA_INICIO_TABLA
    )

    ws: Worksheet = writer.sheets[sheet_name]

    # ── Bloque de identificación institucional ──────────────────────────
    hoy = _dt.date.today().strftime("%d/%m/%Y")
    ficha = [
        ("Nombre de la publicación", "Diccionario de Datos — Exportación SIDOE (Consultas)"),
        ("Institución", "Oficina Nacional de Estadística (ONE), República Dominicana"),
        (
            "Objetivo general",
            "Documentar la estructura, tipo de dato y significado de cada "
            "variable exportada en este archivo, conforme a los "
            "Lineamientos y Recomendaciones para Documentar el Diccionario "
            "de Datos de la ONE.",
        ),
        (
            "Cobertura",
            "Hojas 'Indicadores', 'Fuentes' y 'Factibilidad' de este mismo "
            "archivo, según el filtro de Consultas aplicado al momento de "
            "generarlo.",
        ),
        ("Clasificación", "Diccionario de Datos Pasivo (archivo XLSX)"),
        ("Fecha de generación", hoy),
    ]

    ws.cell(row=1, column=1, value="Diccionario de Datos — SIDOE").font = _FONT_TITULO
    fila = 2
    for etiqueta, valor in ficha:
        c_etq = ws.cell(row=fila, column=1, value=etiqueta)
        c_etq.font = _FONT_ETIQUETA_FICHA
        c_val = ws.cell(row=fila, column=2, value=valor)
        c_val.font = _FONT_NORMAL
        c_val.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=8)
        fila += 1

    # ── Formato de la tabla (encabezados en fila FILA_INICIO_TABLA + 1) ──
    fila_encabezado_tabla = FILA_INICIO_TABLA + 1
    for col_idx in range(1, len(_COLUMNAS_TABLA) + 1):
        celda = ws.cell(row=fila_encabezado_tabla, column=col_idx)
        celda.fill = _FILL_ENCABEZADO
        celda.font = _FONT_ENCABEZADO
        celda.alignment = Alignment(wrap_text=True, vertical="center")

    for col_idx, ancho in enumerate(_ANCHO_COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    for fila_idx in range(fila_encabezado_tabla + 1, fila_encabezado_tabla + 1 + len(tabla)):
        for col_idx in range(1, len(_COLUMNAS_TABLA) + 1):
            ws.cell(row=fila_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        ws.cell(row=fila_idx, column=1).font = Font(bold=True, name="Arial")

    ws.freeze_panes = ws.cell(row=fila_encabezado_tabla + 1, column=1)
