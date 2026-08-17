"""
tests/test_05_integracion_logs_export.py
=========================================
TESTS DE INTEGRACIÓN — Logs de Auditoría, Exportación Excel, Engine vs BD

Validan:
  - registrar_log persiste dentro de transacción y con standalone
  - registrar_log sin usuario_id no lanza excepción (solo warning)
  - generar_excel_memoria devuelve bytes válidos de XLSX
  - El engine recalcula scores consistentes con los valores almacenados en BD
  - Consistencia entre score almacenado y recálculo en tiempo real
"""

import io
import pytest


# ---------------------------------------------------------------------------
# Tests de Logs de Auditoría
# ---------------------------------------------------------------------------

class TestRegistrarLog:

    def test_registrar_log_en_transaccion(self, sidoe_config):
        """registrar_log dentro de una transacción debe persistir al hacer commit."""
        import data.database as db_mod
        from models.logs import registrar_log

        conn = db_mod.obtener_conexion()
        cursor = conn.cursor()
        registrar_log(cursor, usuario_id=1, accion="TEST_LOG", detalle="detalle de prueba")
        conn.commit()

        row = conn.execute(
            "SELECT accion, detalle FROM auditoria WHERE accion='TEST_LOG' "
            "ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert row is not None
        assert row[0] == "TEST_LOG"
        assert row[1] == "detalle de prueba"

    def test_registrar_log_sin_usuario_id_no_lanza_excepcion(self, sidoe_config):
        """Con usuario_id=None el log se omite silenciosamente (solo warning)."""
        import data.database as db_mod
        from models.logs import registrar_log

        conn = db_mod.obtener_conexion()
        cursor = conn.cursor()
        # No debe lanzar excepción
        registrar_log(cursor, usuario_id=None, accion="SIN_USUARIO", detalle="test")
        conn.commit()
        conn.close()

    def test_registrar_log_standalone_persiste(self, sidoe_config):
        """registrar_log_standalone abre su propia conexión y persiste."""
        import data.database as db_mod
        from models.logs import registrar_log_standalone

        registrar_log_standalone(
            usuario_id=1, accion="TEST_STANDALONE", detalle="standalone test"
        )
        conn = db_mod.obtener_conexion()
        row = conn.execute(
            "SELECT accion FROM auditoria WHERE accion='TEST_STANDALONE' "
            "ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        assert row is not None

    def test_log_rollback_no_persiste(self, sidoe_config):
        """Log dentro de un rollback no debe persistir."""
        import data.database as db_mod
        from models.logs import registrar_log

        conn = db_mod.obtener_conexion()
        cursor = conn.cursor()
        registrar_log(cursor, usuario_id=1, accion="LOG_ROLLBACK_TEST", detalle="debe desaparecer")
        conn.rollback()  # Revertir

        row = conn.execute(
            "SELECT accion FROM auditoria WHERE accion='LOG_ROLLBACK_TEST'"
        ).fetchone()
        conn.close()
        assert row is None, "El log no debe persistir tras un rollback"

    def test_auditoria_tiene_timestamp(self, sidoe_config):
        """Los registros de auditoría deben tener una columna de fecha/hora."""
        import data.database as db_mod
        cols = {
            r[1]
            for r in db_mod.obtener_conexion().execute(
                "PRAGMA table_info(auditoria)"
            ).fetchall()
        }
        assert any("fecha" in c.lower() or "timestamp" in c.lower() or "tiempo" in c.lower()
                   for c in cols), "La tabla auditoria no tiene columna de timestamp"


# ---------------------------------------------------------------------------
# Tests de Exportación Excel
# ---------------------------------------------------------------------------

class TestExportExcel:

    def test_generar_excel_devuelve_bytes(self):
        """generar_excel_memoria debe devolver bytes no vacíos."""
        import pandas as pd
        from tracking.export_excel import generar_excel_memoria

        df = pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})
        resultado = generar_excel_memoria(df, "Hoja_Test")
        assert isinstance(resultado, bytes)
        assert len(resultado) > 0

    def test_excel_generado_es_xlsx_valido(self):
        """El archivo generado debe ser un XLSX parseable por openpyxl."""
        import pandas as pd
        import openpyxl
        from tracking.export_excel import generar_excel_memoria

        df = pd.DataFrame({
            "Codigo": ["TEST-001", "TEST-002"],
            "Indicador": ["Indicador A", "Indicador B"],
            "Score": [100.002, 55.0],
        })
        raw = generar_excel_memoria(df, "Matriz_Test")
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert "Matriz_Test" in wb.sheetnames
        ws = wb["Matriz_Test"]
        assert ws.max_row >= 2  # Header + 2 filas

    def test_excel_preserva_todos_los_datos(self):
        """El Excel generado debe contener exactamente los datos del DataFrame."""
        import pandas as pd
        import openpyxl
        from tracking.export_excel import generar_excel_memoria

        registros = [
            {"codigo": f"TEST-{i:03d}", "score": float(i * 10)}
            for i in range(1, 51)
        ]
        df = pd.DataFrame(registros)
        raw = generar_excel_memoria(df, "Test50")
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb["Test50"]
        # Header (fila 1) + 50 filas de datos
        assert ws.max_row == 51

    def test_excel_nombre_hoja_por_defecto(self):
        """Si no se especifica nombre de hoja, usa el valor por defecto."""
        import pandas as pd
        import openpyxl
        from tracking.export_excel import generar_excel_memoria

        df = pd.DataFrame({"X": [1]})
        raw = generar_excel_memoria(df)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert len(wb.sheetnames) >= 1

    def test_excel_con_dataframe_vacio_no_falla(self):
        """Exportar un DataFrame vacío no debe lanzar excepción."""
        import pandas as pd
        from tracking.export_excel import generar_excel_memoria

        df = pd.DataFrame(columns=["A", "B", "C"])
        resultado = generar_excel_memoria(df, "Vacia")
        assert isinstance(resultado, bytes)


# ---------------------------------------------------------------------------
# Tests de consistencia Engine vs datos en BD
# ---------------------------------------------------------------------------

class TestConsistenciaEngineVsBD:

    def test_recalculo_coincide_con_score_almacenado(self, db_conn):
        """Para los primeros 50 indicadores, el recálculo debe coincidir con el score guardado."""
        from features.engine_factibilidad import calcular_reglas_factibilidad

        filas = db_conn.execute("""
            SELECT indicador_id, c1_metodologia, c21_existencia_fuente,
                   c22_disponibilidad, c23_periodicidad_establecida,
                   c31_posee_desagregacion, num_desagregaciones_requeridas,
                   num_desagregaciones_disponibles, articulacion_fuentes,
                   armonizacion_conceptual, subregistro_cobertura,
                   cobertura_territorial, estructura_datos, variables_calculo,
                   score_factibilidad_final, categoria_factibilidad
            FROM calculo_factibilidad
            LIMIT 50
        """).fetchall()

        errores = []
        for fila in filas:
            datos = {
                "c1_metodologia": fila[1],
                "c21_existencia_fuente": fila[2],
                "c22_disponibilidad": fila[3],
                "c23_periodicidad_establecida": fila[4],
                "c31_posee_desagregacion": fila[5],
                "num_desagregaciones_requeridas": fila[6],
                "num_desagregaciones_disponibles": fila[7],
                "articulacion_fuentes": fila[8],
                "armonizacion_conceptual": fila[9],
                "subregistro_cobertura": fila[10],
                "cobertura_territorial": fila[11],
                "estructura_datos": fila[12],
                "variables_calculo": fila[13],
            }
            resultado = calcular_reglas_factibilidad(datos)
            score_bd = round(fila[14], 3) if fila[14] is not None else 0.0
            score_calc = resultado["score_factibilidad_final"]

            if abs(score_bd - score_calc) > 0.01:
                errores.append(
                    f"indicador_id={fila[0]}: BD={score_bd}, Calculado={score_calc}"
                )
            if fila[15] != resultado["categoria_factibilidad"]:
                errores.append(
                    f"indicador_id={fila[0]}: Categoría BD='{fila[15]}', "
                    f"Calculada='{resultado['categoria_factibilidad']}'"
                )

        assert len(errores) == 0, (
            f"{len(errores)} inconsistencias entre BD y Engine:\n"
            + "\n".join(errores[:10])  # Mostrar solo los primeros 10
        )

    def test_categoria_consistente_con_score_en_bd(self, db_conn):
        """Para TODOS los indicadores, la categoría almacenada debe
        ser consistente con el score almacenado y los umbrales de config."""
        from config import UMBRAL_ALTA, UMBRAL_MEDIA, CAT_I, CAT_II, CAT_III

        filas = db_conn.execute(
            "SELECT indicador_id, score_factibilidad_final, categoria_factibilidad "
            "FROM calculo_factibilidad"
        ).fetchall()

        errores = []
        for ind_id, score, categoria in filas:
            score = score or 0.0
            esperada = CAT_I if score >= UMBRAL_ALTA else (CAT_II if score >= UMBRAL_MEDIA else CAT_III)
            if categoria != esperada:
                errores.append(
                    f"indicador_id={ind_id}: score={score}, "
                    f"Categoría almacenada='{categoria}', Esperada='{esperada}'"
                )

        assert len(errores) == 0, (
            f"{len(errores)} inconsistencias de categoría en BD:\n"
            + "\n".join(errores[:10])
        )

    @pytest.mark.requiere_bd_local
    def test_distribucion_factibilidad_coherente(self, db_conn):
        """Verificar que la distribución I/II/III tiene sentido estadístico."""
        dist = db_conn.execute("""
            SELECT categoria_factibilidad, COUNT(*) as n
            FROM calculo_factibilidad
            GROUP BY categoria_factibilidad
        """).fetchall()
        dist_dict = {r[0]: r[1] for r in dist}

        total = sum(dist_dict.values())
        assert total > 800, f"Se esperan > 800 indicadores, hay {total}"

        # Cada categoría debe tener al menos un indicador
        from config import CAT_I, CAT_II, CAT_III
        for cat in [CAT_I, CAT_II, CAT_III]:
            assert cat in dist_dict, f"No hay ningún indicador con categoría '{cat}'"
            assert dist_dict[cat] > 0


# ---------------------------------------------------------------------------
# Tests de helpers de la capa de normalización del engine
# ---------------------------------------------------------------------------

class TestNormalizacionEngine:

    def test_norm_elimina_espacios(self):
        from features.engine_factibilidad import _norm
        assert _norm("  Sí  ") == "Sí"
        assert _norm("  No  ") == "No"

    def test_norm_none_devuelve_vacio(self):
        from features.engine_factibilidad import _norm
        assert _norm(None) == ""

    def test_norm_entero_se_convierte_a_string(self):
        from features.engine_factibilidad import _norm
        assert _norm(42) == "42"
