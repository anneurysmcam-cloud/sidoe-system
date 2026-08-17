"""
tests/test_22_migracion_backfill_contenido_referenciados.py
==============================================================
Cubre data/migracion_backfill_contenido_referenciados.py:

- Cuando el Excel trae el par en AMBAS direcciones, gana el lado con
  mayor score_factibilidad_final -- no el orden de iteración.
- Pares en EXCLUSIONES_CONOCIDAS (o pasados explícitamente) se saltan.
- Empates en score no se propagan automáticamente, se reportan como
  ambiguos.
- Códigos sin match en la BD se reportan, no rompen el resto.
- Reutiliza sincronizar_contenido_referenciados(), así que es idempotente.
"""

import sqlite3

import openpyxl
import pytest

from data.migraciones_historicas.migracion_backfill_contenido_referenciados import migrar
from models.crud_indicadores import guardar_indicador

FACT_ALTA = {
    "c1_metodologia": "Indicador con metodología nacional o internacional definida",
    "c21_existencia_fuente": "Completamente",
    "c22_disponibilidad": "Sí",
    "c23_periodicidad_establecida": "Sí",
    "c31_posee_desagregacion": "Sí",
    "num_desagregaciones_requeridas": 1,
    "num_desagregaciones_disponibles": 1,
    "articulacion_fuentes": "No requiere de articulación",
    "armonizacion_conceptual": "No",
    "subregistro_cobertura": "No",
    "cobertura_territorial": "Sí",
    "estructura_datos": (
        "a) La fuente de información utiliza en el procesamiento "
        "una base de datos estructurada"
    ),
    "variables_calculo": "Sí",
}

FACT_BAJA = {
    "c1_metodologia": "No cumple con los criterios anteriores",
    "c21_existencia_fuente": "No hay fuente",
    "c22_disponibilidad": "No",
    "c23_periodicidad_establecida": "No",
    "c31_posee_desagregacion": "No",
    "num_desagregaciones_requeridas": 0,
    "num_desagregaciones_disponibles": 0,
    "articulacion_fuentes": "No se articula",
    "armonizacion_conceptual": "Sí",
    "subregistro_cobertura": "Sí",
    "cobertura_territorial": "No",
    "estructura_datos": "c) No posee ninguna de las anteriores",
    "variables_calculo": "No",
}


def _crear_excel_prueba(ruta, filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demanda y Oferta"
    ws["A1"] = "DEMANDA"
    ws["F3"] = "Código"
    ws["H3"] = "Indicadores duplicados"
    ws["I3"] = "Indicador"
    fila_actual = 4
    for codigo, dup, nombre in filas:
        ws[f"F{fila_actual}"] = codigo
        ws[f"H{fila_actual}"] = dup
        ws[f"I{fila_actual}"] = nombre
        fila_actual += 1
    wb.save(ruta)


def _indicador_id(conn, codigo):
    return conn.execute("SELECT id FROM indicadores WHERE codigo = ?", (codigo,)).fetchone()[0]


def _sembrar(db_path):
    # ALTA/BAJA: referenciados en AMBAS direcciones en el Excel, con
    # factibilidad claramente distinta -- debe ganar ALTA sin importar
    # el orden de iteración.
    guardar_indicador(
        {"codigo": "ALTA", "indicador": "Indicador con factibilidad alta",
         "estado_indicador": "Activo", "generador_demanda_id": 1},
        [{"nombre_fuente": "Fuente buena", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "BAJA", "indicador": "Indicador con factibilidad baja",
         "estado_indicador": "Activo", "generador_demanda_id": 2},
        [{"nombre_fuente": "Fuente pobre", "institucion_productora": "Vieja"}],
        FACT_BAJA,
    )

    # EMPATE-A/EMPATE-B: misma factibilidad -- debe quedar ambiguo (ninguno
    # de los dos es ODS, así que no hay desempate posible).
    guardar_indicador(
        {"codigo": "EMPATE-A", "indicador": "Indicador empatado A",
         "estado_indicador": "Activo", "generador_demanda_id": 1},  # END
        [{"nombre_fuente": "Fuente A", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "EMPATE-B", "indicador": "Indicador empatado B",
         "estado_indicador": "Activo", "generador_demanda_id": 3},  # CMV
        [{"nombre_fuente": "Fuente B", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )

    # ODS-EMPATE/CMV-EMPATE: misma factibilidad Y contenido distinto, PERO
    # uno de los dos SÍ es ODS -- debe desempatar y ganar el lado ODS.
    guardar_indicador(
        {"codigo": "CMV-EMPATE", "indicador": "Duplicado real, lado CMV",
         "estado_indicador": "Activo", "generador_demanda_id": 3},  # CMV
        [{"nombre_fuente": "Fuente CMV vieja", "institucion_productora": "Vieja"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "ODS-EMPATE", "indicador": "Duplicado real, lado ODS",
         "estado_indicador": "Activo", "generador_demanda_id": 2},  # ODS
        [{"nombre_fuente": "Fuente ODS estándar", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )

    # EXCLUIDO-A/EXCLUIDO-B: se referencian pero deben excluirse explícitamente.
    guardar_indicador(
        {"codigo": "EXCLUIDO-A", "indicador": "Indicador excluido A",
         "estado_indicador": "Activo", "generador_demanda_id": 1},
        [{"nombre_fuente": "Fuente excluida A", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "EXCLUIDO-B", "indicador": "Indicador excluido B",
         "estado_indicador": "Activo", "generador_demanda_id": 2},
        [{"nombre_fuente": "Fuente excluida B", "institucion_productora": "ONE"}],
        FACT_BAJA,
    )

    # Fijar los vínculos bidireccionales tal como los dejaría el backfill de
    # vínculos ya corregido (test_20) -- sincronizar_contenido_referenciados()
    # lee el campo indicadores_duplicados DESDE LA BD del indicador que
    # termina siendo el origen, no recibe el destino por parámetro.
    conn = sqlite3.connect(db_path)
    for a, b in [
        ("ALTA", "BAJA"), ("EMPATE-A", "EMPATE-B"),
        ("EXCLUIDO-A", "EXCLUIDO-B"), ("CMV-EMPATE", "ODS-EMPATE"),
    ]:
        conn.execute("UPDATE indicadores SET indicadores_duplicados = ? WHERE codigo = ?", (b, a))
        conn.execute("UPDATE indicadores SET indicadores_duplicados = ? WHERE codigo = ?", (a, b))
    conn.commit()
    conn.close()


@pytest.fixture
def excel_prueba(tmp_path):
    ruta = tmp_path / "excel_contenido_prueba.xlsx"
    _crear_excel_prueba(ruta, [
        ("ALTA", "BAJA", "Indicador con factibilidad alta"),
        ("BAJA", "ALTA", "Indicador con factibilidad baja"),
        ("EMPATE-A", "EMPATE-B", "Indicador empatado A"),
        ("EMPATE-B", "EMPATE-A", "Indicador empatado B"),
        ("CMV-EMPATE", "ODS-EMPATE", "Duplicado real, lado CMV"),
        ("ODS-EMPATE", "CMV-EMPATE", "Duplicado real, lado ODS"),
        ("EXCLUIDO-A", "EXCLUIDO-B", "Indicador excluido A"),
        ("HUERFANO", "NO-EXISTE-999", "Indicador con destino inexistente"),
    ])
    return str(ruta)


class TestBackfillContenidoGanaMayorFactibilidad:

    def test_gana_el_de_mayor_factibilidad_sin_importar_direccion_excel(
        self, sidoe_config, excel_prueba
    ):
        _sembrar(sidoe_config)

        resumen = migrar(
            excel_prueba, db_path=sidoe_config,
            exclusiones=frozenset({frozenset({"EXCLUIDO-A", "EXCLUIDO-B"})}),
        )

        propagacion = next(
            p for p in resumen["propagados"]
            if {p["origen"]} | set(p["destinos"]) == {"ALTA", "BAJA"}
        )
        assert propagacion["origen"] == "ALTA"

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fuente_baja = conn.execute(
            "SELECT nombre_fuente FROM fuentes_indicador WHERE indicador_id = "
            f"{_indicador_id(conn, 'BAJA')}"
        ).fetchone()
        assert fuente_baja["nombre_fuente"] == "Fuente buena"
        conn.close()

    def test_empate_no_propaga_y_se_reporta_como_ambiguo(self, sidoe_config, excel_prueba):
        _sembrar(sidoe_config)

        resumen = migrar(
            excel_prueba, db_path=sidoe_config,
            exclusiones=frozenset({frozenset({"EXCLUIDO-A", "EXCLUIDO-B"})}),
        )

        assert sorted(resumen["ambiguos"][0]) == ["EMPATE-A", "EMPATE-B"]

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fuente_b = conn.execute(
            "SELECT nombre_fuente FROM fuentes_indicador WHERE indicador_id = "
            f"{_indicador_id(conn, 'EMPATE-B')}"
        ).fetchone()
        assert fuente_b["nombre_fuente"] == "Fuente B"  # no se tocó
        conn.close()

    def test_empate_con_lado_ods_desempata_a_favor_de_ods(self, sidoe_config, excel_prueba):
        _sembrar(sidoe_config)

        resumen = migrar(
            excel_prueba, db_path=sidoe_config,
            exclusiones=frozenset({frozenset({"EXCLUIDO-A", "EXCLUIDO-B"})}),
        )

        propagacion = next(
            p for p in resumen["propagados"]
            if {p["origen"]} | set(p["destinos"]) == {"CMV-EMPATE", "ODS-EMPATE"}
        )
        assert propagacion["origen"] == "ODS-EMPATE"
        assert propagacion["motivo"] == "empate_desempatado_por_ods"

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fuente_cmv = conn.execute(
            "SELECT nombre_fuente FROM fuentes_indicador WHERE indicador_id = "
            f"{_indicador_id(conn, 'CMV-EMPATE')}"
        ).fetchone()
        assert fuente_cmv["nombre_fuente"] == "Fuente ODS estándar"
        conn.close()

    def test_par_excluido_no_se_toca(self, sidoe_config, excel_prueba):
        _sembrar(sidoe_config)

        resumen = migrar(
            excel_prueba, db_path=sidoe_config,
            exclusiones=frozenset({frozenset({"EXCLUIDO-A", "EXCLUIDO-B"})}),
        )

        assert sorted(resumen["excluidos"][0]) == ["EXCLUIDO-A", "EXCLUIDO-B"]

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fuente_b = conn.execute(
            "SELECT nombre_fuente FROM fuentes_indicador WHERE indicador_id = "
            f"{_indicador_id(conn, 'EXCLUIDO-B')}"
        ).fetchone()
        assert fuente_b["nombre_fuente"] == "Fuente excluida B"  # sin cambios
        conn.close()

    def test_par_sin_match_se_reporta(self, sidoe_config, excel_prueba):
        _sembrar(sidoe_config)

        resumen = migrar(
            excel_prueba, db_path=sidoe_config,
            exclusiones=frozenset({frozenset({"EXCLUIDO-A", "EXCLUIDO-B"})}),
        )

        assert any("NO-EXISTE-999" in par for par in resumen["sin_match"])

    def test_es_idempotente(self, sidoe_config, excel_prueba):
        _sembrar(sidoe_config)
        exclusiones = frozenset({frozenset({"EXCLUIDO-A", "EXCLUIDO-B"})})

        primero = migrar(excel_prueba, db_path=sidoe_config, exclusiones=exclusiones)
        segundo = migrar(excel_prueba, db_path=sidoe_config, exclusiones=exclusiones)

        assert primero["propagados"] != []
        assert segundo["propagados"] == []
        assert any(
            sorted(par) == ["ALTA", "BAJA"] for par in segundo["ya_al_dia"]
        ), "Tras sincronizar, la segunda corrida debe reconocerlo como 'ya al día', no como ambiguo"

    def test_excel_inexistente_no_rompe(self, sidoe_config, tmp_path):
        _sembrar(sidoe_config)

        resultado = migrar(str(tmp_path / "no_existe.xlsx"), db_path=sidoe_config)

        assert resultado == {"error": "excel_no_encontrado"}
