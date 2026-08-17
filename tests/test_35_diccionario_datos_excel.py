"""
tests/test_35_diccionario_datos_excel.py
=========================================
TEST DE REGRESIÓN — hoja "Diccionario de Datos" en el Excel de Consultas.

Contexto: se agregó data/diccionario_datos.py para documentar, dentro del
propio Excel exportado desde Consultas, cada columna de las hojas
"Indicadores", "Fuentes" y "Factibilidad", conforme a los Lineamientos y
Recomendaciones para Documentar el Diccionario de Datos de la ONE
(Diccionario de Datos Pasivo — archivo XLSX).

Verifica:
  1. La tabla incluye una fila por cada columna realmente presente en cada
     DataFrame exportado (no una lista fija hardcodeada).
  2. Las columnas conocidas del vocabulario fijo tienen su metadato
     documentado (no caen en el fallback genérico).
  3. Un campo dinámico de Auxiliares (no está en el catálogo estático) cae
     en el fallback genérico y aun así queda documentado, no se omite.
  4. El archivo Excel final generado por consultas.py incluye la hoja
     "Diccionario de Datos" junto a las otras tres.
"""

import io

import pandas as pd

from data.diccionario_datos import (
    construir_tabla_diccionario_datos,
    escribir_hoja_diccionario_datos,
)


def _dfs_de_prueba():
    df_ind = pd.DataFrame({
        "codigo": ["IND-001"],
        "generador_demanda": ["END"],
        "indicador": ["Indicador de prueba"],
        "num_fuentes": [2],
        "categoria_personalizada_x": ["valor libre"],  # simula Auxiliares
    })
    df_fuentes = pd.DataFrame({
        "indicador_codigo": ["IND-001"],
        "indicador_nombre": ["Indicador de prueba"],
        "nombre_fuente": ["Encuesta de prueba"],
        "comentarios": [""],
    })
    df_fac = pd.DataFrame({
        "codigo": ["IND-001"],
        "indicador": ["Indicador de prueba"],
        "generador_demanda": ["END"],
        "puntaje": [95.0],
        "factibilidad": ["Factibilidad I"],
        "calc_timestamp": ["2026-08-11 10:00:00"],
    })
    return df_ind, df_fuentes, df_fac


class TestConstruirTablaDiccionarioDatos:

    def test_una_fila_por_columna_realmente_exportada(self):
        df_ind, df_fuentes, df_fac = _dfs_de_prueba()
        tabla = construir_tabla_diccionario_datos(df_ind, df_fuentes, df_fac)

        total_columnas_esperado = len(df_ind.columns) + len(df_fuentes.columns) + len(df_fac.columns)
        assert len(tabla) == total_columnas_esperado

        # Cada columna de cada hoja aparece documentada bajo su propia "Hoja".
        for hoja, df in (("Indicadores", df_ind), ("Fuentes", df_fuentes), ("Factibilidad", df_fac)):
            nombres_documentados = set(
                tabla.loc[tabla["Hoja"] == hoja, "Nombre de la Variable"]
            )
            assert nombres_documentados == set(df.columns)

    def test_columnas_del_vocabulario_fijo_no_usan_fallback_generico(self):
        df_ind, df_fuentes, df_fac = _dfs_de_prueba()
        tabla = construir_tabla_diccionario_datos(df_ind, df_fuentes, df_fac)

        fila_codigo = tabla[
            (tabla["Hoja"] == "Indicadores") & (tabla["Nombre de la Variable"] == "codigo")
        ].iloc[0]
        assert fila_codigo["Etiqueta"] == "Código del Indicador"
        assert "personalizado" not in fila_codigo["Descripción"].lower()

        fila_factibilidad = tabla[
            (tabla["Hoja"] == "Factibilidad") & (tabla["Nombre de la Variable"] == "factibilidad")
        ].iloc[0]
        assert "Factibilidad I" in fila_factibilidad["Dominio / Valores Permitidos"]

    def test_campo_dinamico_de_auxiliares_cae_en_fallback_pero_queda_documentado(self):
        df_ind, df_fuentes, df_fac = _dfs_de_prueba()
        tabla = construir_tabla_diccionario_datos(df_ind, df_fuentes, df_fac)

        fila_custom = tabla[
            (tabla["Hoja"] == "Indicadores")
            & (tabla["Nombre de la Variable"] == "categoria_personalizada_x")
        ].iloc[0]
        assert fila_custom["Etiqueta"] == "Categoria Personalizada X"
        assert "Auxiliares" in fila_custom["Dominio / Valores Permitidos"]
        assert fila_custom["Descripción"]  # no vacío: sigue documentado


class TestEscribirHojaEnExcel:

    def test_hoja_diccionario_de_datos_se_agrega_al_excel(self):
        df_ind, df_fuentes, df_fac = _dfs_de_prueba()

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_ind.to_excel(writer, index=False, sheet_name="Indicadores")
            df_fuentes.to_excel(writer, index=False, sheet_name="Fuentes")
            df_fac.to_excel(writer, index=False, sheet_name="Factibilidad")
            escribir_hoja_diccionario_datos(writer, df_ind, df_fuentes, df_fac)

        buf.seek(0)
        libro = pd.ExcelFile(buf, engine="openpyxl")
        assert libro.sheet_names == [
            "Indicadores", "Fuentes", "Factibilidad", "Diccionario de Datos",
        ]

        hoja_dd = pd.read_excel(buf, sheet_name="Diccionario de Datos", header=None)
        # El bloque de ficha técnica ocupa las primeras filas antes de la tabla.
        assert "Diccionario de Datos" in str(hoja_dd.iloc[0, 0])
        assert "Nombre de la publicación" in str(hoja_dd.iloc[1, 0])


class TestFormatoEncabezadoHojasDatos:
    """A pedido de Randy: los encabezados de las hojas 'Indicadores',
    'Fuentes' y 'Factibilidad' del Excel exportado deben usar el mismo
    estilo azul/blanco institucional que ya tenía la hoja 'Diccionario de
    Datos' — ver data.diccionario_datos.aplicar_formato_encabezado_hoja_datos
    y su uso en views/consultas.py."""

    def test_encabezado_de_las_tres_hojas_usa_el_mismo_estilo_que_diccionario(self):
        from data.diccionario_datos import (
            _FILL_ENCABEZADO,
            _FONT_ENCABEZADO,
            aplicar_formato_encabezado_hoja_datos,
        )

        df_ind, df_fuentes, df_fac = _dfs_de_prueba()

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_ind.to_excel(writer, index=False, sheet_name="Indicadores")
            df_fuentes.to_excel(writer, index=False, sheet_name="Fuentes")
            df_fac.to_excel(writer, index=False, sheet_name="Factibilidad")
            aplicar_formato_encabezado_hoja_datos(
                writer.sheets["Indicadores"], len(df_ind.columns)
            )
            aplicar_formato_encabezado_hoja_datos(
                writer.sheets["Fuentes"], len(df_fuentes.columns)
            )
            aplicar_formato_encabezado_hoja_datos(
                writer.sheets["Factibilidad"], len(df_fac.columns)
            )
            escribir_hoja_diccionario_datos(writer, df_ind, df_fuentes, df_fac)

        buf.seek(0)
        from openpyxl import load_workbook
        libro = load_workbook(buf)

        for nombre_hoja, df in (
            ("Indicadores", df_ind), ("Fuentes", df_fuentes), ("Factibilidad", df_fac),
        ):
            ws = libro[nombre_hoja]
            for col_idx in range(1, len(df.columns) + 1):
                celda = ws.cell(row=1, column=col_idx)
                assert celda.fill.fgColor.rgb == _FILL_ENCABEZADO.fgColor.rgb, (
                    f"Encabezado de '{nombre_hoja}' col {col_idx} no tiene "
                    "el fondo azul institucional."
                )
                assert celda.font.bold is True
                assert celda.font.color.rgb == _FONT_ENCABEZADO.color.rgb, (
                    f"Encabezado de '{nombre_hoja}' col {col_idx} no tiene "
                    "el texto en blanco."
                )
            # La fila de DATOS (fila 2) no debe heredar el estilo del
            # encabezado — solo la fila 1.
            celda_dato = ws.cell(row=2, column=1)
            assert celda_dato.fill.fgColor.rgb != _FILL_ENCABEZADO.fgColor.rgb

    def test_hoja_vacia_no_revienta_el_formato(self):
        """Un DataFrame sin filas (filtro de Consultas sin resultados)
        sigue teniendo columnas — el formato de encabezado debe aplicarse
        igual sin lanzar excepciones."""
        from data.diccionario_datos import aplicar_formato_encabezado_hoja_datos

        df_vacio = pd.DataFrame(columns=["codigo", "indicador"])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_vacio.to_excel(writer, index=False, sheet_name="Indicadores")
            aplicar_formato_encabezado_hoja_datos(
                writer.sheets["Indicadores"], len(df_vacio.columns)
            )
        # No debe lanzar excepción; el archivo debe seguir siendo válido.
        buf.seek(0)
        libro = pd.ExcelFile(buf, engine="openpyxl")
        assert "Indicadores" in libro.sheet_names


class TestComentariosPuntajeFactibilidad:
    """A pedido de Randy: los campos de 'Factibilidad' que tienen una
    puntuación en el Engine (features/engine_factibilidad.py) deben mostrar
    el puntaje de cada opción posible en 'Comentarios Adicionales' del
    Diccionario de Datos."""

    def test_campos_con_puntaje_documentan_valor_por_opcion(self):
        df_fac = pd.DataFrame({
            "codigo": ["IND-001"],
            "generador_demanda": ["END"],
            "c1_metodologia": ["Indicador con metodología nacional o "
                                "internacional definida"],
        })
        tabla = construir_tabla_diccionario_datos(
            pd.DataFrame({"codigo": ["IND-001"]}), pd.DataFrame(), df_fac
        )

        filas_fac = tabla[tabla["Hoja"] == "Factibilidad"].set_index(
            "Nombre de la Variable"
        )
        assert "15" in filas_fac.loc["c1_metodologia", "Comentarios Adicionales"]
        assert filas_fac.loc["generador_demanda", "Comentarios Adicionales"] == ""

    def test_todos_los_criterios_del_engine_tienen_comentario_de_puntaje(self):
        """Cada criterio del Engine (C1 a Variables de Cálculo) debe tener
        su puntaje documentado; ninguno debe quedar con comentario vacío."""
        from data.diccionario_datos import _META_FACTIBILIDAD

        campos_con_puntaje = [
            "c1_metodologia", "c21_existencia_fuente", "c22_disponibilidad",
            "c23_periodicidad_establecida", "c31_posee_desagregacion",
            "num_desagregaciones_requeridas", "num_desagregaciones_disponibles",
            "articulacion_fuentes", "armonizacion_conceptual",
            "subregistro_cobertura", "cobertura_territorial",
            "estructura_datos", "variables_calculo",
        ]
        for campo in campos_con_puntaje:
            comentario = _META_FACTIBILIDAD[campo].get("comentarios", "")
            assert comentario, f"'{campo}' no tiene comentario de puntaje documentado"


class TestAjustarAnchoColumnasAuto:
    """A pedido de Randy: el ancho de columna del Excel exportado debe
    ajustarse al contenido real (encabezado + valores) para que no se vea
    texto cortado, en vez del ancho fijo por defecto de openpyxl."""

    def test_columnas_con_contenido_largo_quedan_mas_anchas_que_las_cortas(self):
        from openpyxl.utils import get_column_letter

        from data.diccionario_datos import ajustar_ancho_columnas_auto

        df = pd.DataFrame({
            "codigo": ["1.1", "1.2"],
            "metodo_calculo": [
                "Este es un método de cálculo con una descripción bastante "
                "larga que en el ancho por defecto se vería cortado",
                "Otro método de cálculo igual de largo para el promedio",
            ],
        })
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Indicadores")
            ws = writer.sheets["Indicadores"]
            ajustar_ancho_columnas_auto(ws, df)

            ancho_codigo = ws.column_dimensions[get_column_letter(1)].width
            ancho_metodo = ws.column_dimensions[get_column_letter(2)].width
            assert ancho_metodo > ancho_codigo
            # Columna corta respeta el ancho mínimo por legibilidad.
            assert ancho_codigo >= 10
            # Columna larga queda acotada al ancho máximo (no desborda la hoja).
            assert ancho_metodo <= 45

    def test_encabezado_largo_no_queda_acotado_por_el_ancho_maximo(self):
        """Pedido explícito de Randy: los encabezados deben caber siempre
        en una sola línea, aunque el nombre de columna sea largo (a
        diferencia del contenido de datos, que sí se acota a ancho_max)."""
        from openpyxl.utils import get_column_letter

        from data.diccionario_datos import ajustar_ancho_columnas_auto

        columna_larga = "num_desagregaciones_disponibles"  # 32 caracteres
        df = pd.DataFrame({columna_larga: [1, 2]})
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Indicadores")
            ws = writer.sheets["Indicadores"]
            ajustar_ancho_columnas_auto(ws, df, ancho_max=20)
            ancho = ws.column_dimensions[get_column_letter(1)].width
            assert ancho >= len(columna_larga)

    def test_valores_nulos_no_revientan_el_ajuste_de_ancho(self):
        """Regresión: TypeError: object of type 'float' has no len() —
        con pandas 3.x (backend Arrow), astype(str).map(len) fallaba en
        columnas con valores nulos (NaN/None), tal como ocurría al
        exportar indicadores/fuentes reales con campos opcionales vacíos."""
        from data.diccionario_datos import ajustar_ancho_columnas_auto

        df = pd.DataFrame({
            "codigo": ["1.1", "1.2"],
            "especificar_clasificacion": [None, "Texto largo de ejemplo"],
            "num_desagregaciones_requeridas": [1, None],
        })
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Indicadores")
            ws = writer.sheets["Indicadores"]
            ajustar_ancho_columnas_auto(ws, df)  # no debe lanzar excepción

    def test_hoja_vacia_no_revienta_el_ajuste_de_ancho(self):
        from data.diccionario_datos import ajustar_ancho_columnas_auto

        df_vacio = pd.DataFrame(columns=["codigo", "indicador"])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_vacio.to_excel(writer, index=False, sheet_name="Indicadores")
            ajustar_ancho_columnas_auto(writer.sheets["Indicadores"], df_vacio)
        buf.seek(0)
        libro = pd.ExcelFile(buf, engine="openpyxl")
        assert "Indicadores" in libro.sheet_names
