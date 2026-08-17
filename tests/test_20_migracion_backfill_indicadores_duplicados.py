"""
tests/test_20_migracion_backfill_indicadores_duplicados.py
============================================================
Cubre data/migracion_backfill_indicadores_duplicados.py:

- Backfill correcto cruzando por código (incluye agrupación por código
  cuando el indicador tiene varias filas de fuente en el Excel).
- Resuelve el prefijo de generador de la columna "Indicadores duplicados"
  (ej. "CMV A.1" -> codigo real "A.1"), incluyendo variantes con texto
  descriptivo extra (ej. "Componente ODS 11.a.1" -> "11.a.1"). [BUG]
  La primera versión comparaba el string crudo contra `codigo` y nunca
  encontraba match porque cada generador usa códigos "pelados" en su
  propia columna Código (CMV: A.1, ODS: 1.1.1, END: 1.1).
- Referencias no resolubles a un indicador puntual (ej. "PNPSP" sin
  número, ya que PNPSP no tiene códigos individuales en el Excel) se
  reportan aparte, no como "sin match" ni como error.
- Sincronización bidireccional real vía sincronizar_indicadores_referenciados()
  (el otro lado del vínculo también queda actualizado).
- Códigos del Excel que no existen en la BD se reportan, no rompen el run.
- Idempotencia: correr dos veces no duplica trabajo ni cambia el resultado.
"""

import sqlite3

import openpyxl
import pytest

from data.migraciones_historicas.migracion_backfill_indicadores_duplicados import (
    leer_referencias_excel,
    migrar,
)


def _crear_excel_prueba(ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demanda y Oferta"
    # Header real en la fila 3 (header=2 en pandas), igual que el Excel oficial.
    ws["A1"] = "DEMANDA"
    encabezados = {
        "F3": "Código",
        "H3": "Indicadores duplicados",
        "I3": "Indicador",
    }
    for celda, valor in encabezados.items():
        ws[celda] = valor

    filas = [
        # (Código, Indicadores duplicados, Indicador) — dos filas de fuente
        # para el mismo código, valor consistente, debe colapsar a 1.
        # El valor trae el prefijo de generador "CMV", como en el Excel real
        # -- el codigo destino resuelto debe ser el "pelado" "A.1".
        ("1.1.1", "CMV A.1", "Indicador A"),
        ("1.1.1", "CMV A.1", "Indicador A"),
        # Referencia con texto descriptivo extra antes del prefijo, como
        # ocurre en el Excel real ("Componente ODS 11.a.1").
        ("A.24", "Componente ODS 11.a.1", "Indicador CMV con referencia descriptiva"),
        # Referencia genérica a PNPSP, sin número -- no resoluble a un
        # indicador puntual (PNPSP no tiene códigos individuales).
        ("1.2", "PNPSP", "Indicador con referencia no resoluble"),
        # Código sin referencia cruzada — debe ignorarse.
        ("2.2.2", None, "Indicador C sin referencia"),
        # Código que no existe en la BD — debe reportarse como sin match.
        ("9.9.9", "ZZZ-999", "Indicador fantasma"),
    ]
    fila_actual = 4
    for codigo, dup, nombre in filas:
        ws[f"F{fila_actual}"] = codigo
        ws[f"H{fila_actual}"] = dup
        ws[f"I{fila_actual}"] = nombre
        fila_actual += 1

    wb.save(ruta)


def _sembrar_indicadores(db_path):
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute(
        "INSERT INTO indicadores (codigo, indicador, estado_indicador) "
        "VALUES ('1.1.1', 'Indicador A', 'Activo')"
    )
    # El destino real en la BD usa el codigo "pelado" del generador CMV,
    # SIN el prefijo "CMV " que trae la columna del Excel.
    conn.execute(
        "INSERT INTO indicadores (codigo, indicador, estado_indicador) "
        "VALUES ('A.1', 'Indicador B (destino de la referencia)', 'Activo')"
    )
    conn.execute(
        "INSERT INTO indicadores (codigo, indicador, estado_indicador) "
        "VALUES ('A.24', 'Indicador CMV con referencia descriptiva', 'Activo')"
    )
    conn.execute(
        "INSERT INTO indicadores (codigo, indicador, estado_indicador) "
        "VALUES ('11.a.1', 'Indicador ODS destino de referencia descriptiva', 'Activo')"
    )
    conn.execute(
        "INSERT INTO indicadores (codigo, indicador, estado_indicador) "
        "VALUES ('1.2', 'Indicador con referencia no resoluble', 'Activo')"
    )
    conn.execute(
        "INSERT INTO indicadores (codigo, indicador, estado_indicador) "
        "VALUES ('2.2.2', 'Indicador C sin referencia', 'Activo')"
    )
    conn.commit()
    conn.close()


@pytest.fixture
def excel_prueba(tmp_path):
    ruta = tmp_path / "excel_backfill_prueba.xlsx"
    _crear_excel_prueba(ruta)
    return str(ruta)


class TestLeerReferenciasExcelResuelveCodigoDestino:
    """[BUG] "Indicadores duplicados" trae el prefijo del generador
    (ej. "CMV A.1"), no el `codigo` pelado real ("A.1"). Cubre la
    resolución del código destino directamente, sin pasar por la BD."""

    def test_resuelve_prefijo_de_generador_simple(self, excel_prueba):
        referencias = leer_referencias_excel(excel_prueba)
        assert referencias["1.1.1"] == "A.1"

    def test_resuelve_prefijo_con_texto_descriptivo_extra(self, excel_prueba):
        referencias = leer_referencias_excel(excel_prueba)
        assert referencias["A.24"] == "11.a.1"

    def test_referencia_pnpsp_sin_numero_no_es_resoluble(self, excel_prueba):
        referencias = leer_referencias_excel(excel_prueba)
        assert "1.2" not in referencias


class TestBackfillIndicadoresDuplicados:

    def test_backfill_mapea_por_codigo_y_agrupa_filas_repetidas(
        self, sidoe_config, excel_prueba
    ):
        _sembrar_indicadores(sidoe_config)

        resumen = migrar(excel_prueba, db_path=sidoe_config)

        # 1.1.1 -> A.1, A.24 -> 11.a.1 (2 actualizados); 1.2 (PNPSP) no
        # cuenta porque leer_referencias_excel ya lo excluye.
        assert resumen["actualizados"] == 2
        assert resumen["sin_match"] == ["9.9.9"]

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fila = conn.execute(
            "SELECT indicadores_duplicados FROM indicadores WHERE codigo = '1.1.1'"
        ).fetchone()
        assert fila["indicadores_duplicados"] == "A.1"
        conn.close()

    def test_backfill_sincroniza_bidireccionalmente(self, sidoe_config, excel_prueba):
        """El indicador REFERENCIADO (A.1) también debe apuntar de vuelta,
        vía sincronizar_indicadores_referenciados() — no solo el que trae el
        valor en el Excel."""
        _sembrar_indicadores(sidoe_config)

        migrar(excel_prueba, db_path=sidoe_config)

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fila = conn.execute(
            "SELECT indicadores_duplicados FROM indicadores WHERE codigo = 'A.1'"
        ).fetchone()
        assert fila["indicadores_duplicados"] == "1.1.1"
        conn.close()

    def test_backfill_resuelve_referencia_con_texto_descriptivo(self, sidoe_config, excel_prueba):
        _sembrar_indicadores(sidoe_config)

        migrar(excel_prueba, db_path=sidoe_config)

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        origen = conn.execute(
            "SELECT indicadores_duplicados FROM indicadores WHERE codigo = 'A.24'"
        ).fetchone()
        destino = conn.execute(
            "SELECT indicadores_duplicados FROM indicadores WHERE codigo = '11.a.1'"
        ).fetchone()
        assert origen["indicadores_duplicados"] == "11.a.1"
        assert destino["indicadores_duplicados"] == "A.24"
        conn.close()

    def test_codigo_sin_referencia_cruzada_no_se_toca(self, sidoe_config, excel_prueba):
        _sembrar_indicadores(sidoe_config)

        migrar(excel_prueba, db_path=sidoe_config)

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fila = conn.execute(
            "SELECT indicadores_duplicados FROM indicadores WHERE codigo = '2.2.2'"
        ).fetchone()
        assert fila["indicadores_duplicados"] is None
        conn.close()

    def test_referencia_pnpsp_no_resoluble_no_toca_el_indicador(self, sidoe_config, excel_prueba):
        _sembrar_indicadores(sidoe_config)

        migrar(excel_prueba, db_path=sidoe_config)

        conn = sqlite3.connect(sidoe_config)
        conn.row_factory = sqlite3.Row
        fila = conn.execute(
            "SELECT indicadores_duplicados FROM indicadores WHERE codigo = '1.2'"
        ).fetchone()
        assert fila["indicadores_duplicados"] is None
        conn.close()

    def test_backfill_es_idempotente(self, sidoe_config, excel_prueba):
        _sembrar_indicadores(sidoe_config)

        primero = migrar(excel_prueba, db_path=sidoe_config)
        segundo = migrar(excel_prueba, db_path=sidoe_config)

        assert primero["actualizados"] == 2
        assert segundo["actualizados"] == 0
        assert segundo["ya_al_dia"] == 2

    def test_excel_inexistente_no_rompe(self, sidoe_config, tmp_path):
        _sembrar_indicadores(sidoe_config)

        resultado = migrar(str(tmp_path / "no_existe.xlsx"), db_path=sidoe_config)

        assert resultado == {"error": "excel_no_encontrado"}
