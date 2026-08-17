"""
tests/test_23_reporte_antes_despues_contenido_referenciados.py
=================================================================
Cubre data/reporte_antes_despues_contenido_referenciados.py:

- Es de SOLO LECTURA: no modifica ni fuentes ni factibilidad en la BD.
- Usa la MISMA lógica que el backfill real: gana el de mayor
  score_factibilidad_final, pares únicos sin dirección, exclusiones.
- Marca correctamente "fuentes_cambian"/"criterios_cambian" para el lado
  que pierde, y "NO" cuando ya coinciden.
- Pares excluidos, sin match, o empatados se reportan con su estado
  explícito, sin romper el resto del reporte.
"""

import csv
import sqlite3

import openpyxl
import pytest

from data.migraciones_historicas.reporte_antes_despues_contenido_referenciados import (
    generar_reporte,
)
from models.crud_indicadores import guardar_indicador

FACT_ALTA = {
    "c1_metodologia": "Indicador con metodología nacional o internacional definida",
    "c21_existencia_fuente": "Completamente",
    "c22_disponibilidad": "Sí",
    "c23_periodicidad_establecida": "Sí",
    "c31_posee_desagregacion": "Sí",
    "num_desagregaciones_requeridas": 1,
    "num_desagregaciones_disponibles": 1,
    "articulacion_fuentes": "No requiere de articulación",
    "armonizacion_conceptual": "No",
    "subregistro_cobertura": "No",
    "cobertura_territorial": "Sí",
    "estructura_datos": (
        "a) La fuente de información utiliza en el procesamiento "
        "una base de datos estructurada"
    ),
    "variables_calculo": "Sí",
}

FACT_BAJA = {
    "c1_metodologia": "No cumple con los criterios anteriores",
    "c21_existencia_fuente": "No hay fuente",
    "c22_disponibilidad": "No",
    "c23_periodicidad_establecida": "No",
    "c31_posee_desagregacion": "No",
    "num_desagregaciones_requeridas": 0,
    "num_desagregaciones_disponibles": 0,
    "articulacion_fuentes": "No se articula",
    "armonizacion_conceptual": "Sí",
    "subregistro_cobertura": "Sí",
    "cobertura_territorial": "No",
    "estructura_datos": "c) No posee ninguna de las anteriores",
    "variables_calculo": "No",
}


def _crear_excel_prueba(ruta, filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demanda y Oferta"
    ws["A1"] = "DEMANDA"
    ws["F3"] = "Código"
    ws["H3"] = "Indicadores duplicados"
    ws["I3"] = "Indicador"
    fila_actual = 4
    for codigo, dup, nombre in filas:
        ws[f"F{fila_actual}"] = codigo
        ws[f"H{fila_actual}"] = dup
        ws[f"I{fila_actual}"] = nombre
        fila_actual += 1
    wb.save(ruta)


@pytest.fixture
def excel_prueba(tmp_path):
    ruta = tmp_path / "excel_reporte_prueba.xlsx"
    _crear_excel_prueba(ruta, [
        ("ORIG-A", "DEST-A", "Indicador con factibilidad alta"),
        ("DEST-A", "ORIG-A", "Indicador con factibilidad baja"),
        ("ORIG-B", "DEST-B", "Indicador origen sin diferencias"),
        ("ORIG-C", "NO-EXISTE-999", "Indicador con destino inexistente"),
        ("CMV-EMPATE", "ODS-EMPATE", "Duplicado real, lado CMV"),
        ("ODS-EMPATE", "CMV-EMPATE", "Duplicado real, lado ODS"),
    ])
    return str(ruta)


def _sembrar(sidoe_config):
    # Par A: factibilidad claramente distinta -> ORIG-A debe ganar.
    guardar_indicador(
        {"codigo": "ORIG-A", "indicador": "Origen A", "estado_indicador": "Activo", "generador_demanda_id": 1},
        [{"nombre_fuente": "Fuente A", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "DEST-A", "indicador": "Destino A, título distinto", "estado_indicador": "Activo", "generador_demanda_id": 2},
        [{"nombre_fuente": "Fuente vieja de destino A", "institucion_productora": "Vieja"}],
        FACT_BAJA,
    )

    # Par B: ya con la misma fuente/factibilidad -> sin cambios.
    guardar_indicador(
        {"codigo": "ORIG-B", "indicador": "Origen B", "estado_indicador": "Activo", "generador_demanda_id": 1},
        [{"nombre_fuente": "Fuente idéntica", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "DEST-B", "indicador": "Destino B, título distinto", "estado_indicador": "Activo", "generador_demanda_id": 2},
        [{"nombre_fuente": "Fuente idéntica", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )

    guardar_indicador(
        {"codigo": "ORIG-C", "indicador": "Origen C sin destino", "estado_indicador": "Activo", "generador_demanda_id": 1},
        [{"nombre_fuente": "Fuente C", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )

    # Empate real con un lado ODS -- debe desempatar a favor de ODS.
    guardar_indicador(
        {"codigo": "CMV-EMPATE", "indicador": "Duplicado real, lado CMV",
         "estado_indicador": "Activo", "generador_demanda_id": 3},  # CMV
        [{"nombre_fuente": "Fuente CMV vieja", "institucion_productora": "Vieja"}],
        FACT_ALTA,
    )
    guardar_indicador(
        {"codigo": "ODS-EMPATE", "indicador": "Duplicado real, lado ODS",
         "estado_indicador": "Activo", "generador_demanda_id": 2},  # ODS
        [{"nombre_fuente": "Fuente ODS estándar", "institucion_productora": "ONE"}],
        FACT_ALTA,
    )

    conn = sqlite3.connect(sidoe_config)
    conn.execute("UPDATE indicadores SET indicadores_duplicados = NULL")
    conn.commit()
    conn.close()


class TestReporteAntesDespues:

    def test_reporte_es_de_solo_lectura(self, sidoe_config, excel_prueba, tmp_path):
        _sembrar(sidoe_config)

        conn = sqlite3.connect(sidoe_config)
        antes = conn.execute(
            "SELECT nombre_fuente FROM fuentes_indicador WHERE indicador_id = "
            "(SELECT id FROM indicadores WHERE codigo = 'DEST-A')"
        ).fetchone()[0]
        conn.close()

        salida = tmp_path / "reporte.csv"
        generar_reporte(excel_prueba, str(salida), db_path=sidoe_config)

        conn = sqlite3.connect(sidoe_config)
        despues = conn.execute(
            "SELECT nombre_fuente FROM fuentes_indicador WHERE indicador_id = "
            "(SELECT id FROM indicadores WHERE codigo = 'DEST-A')"
        ).fetchone()[0]
        conn.close()

        assert antes == despues == "Fuente vieja de destino A", (
            "El reporte NO debe modificar la BD"
        )

    def test_gana_el_de_mayor_factibilidad_no_el_orden_del_excel(
        self, sidoe_config, excel_prueba, tmp_path
    ):
        _sembrar(sidoe_config)
        salida = tmp_path / "reporte.csv"
        generar_reporte(excel_prueba, str(salida), db_path=sidoe_config)

        with open(salida, encoding="utf-8-sig") as f:
            filas = list(csv.DictReader(f))

        fila_par_a = next(
            r for r in filas if {r["codigo_gana"], r["codigo_pierde"]} == {"ORIG-A", "DEST-A"}
        )
        assert fila_par_a["codigo_gana"] == "ORIG-A"
        assert fila_par_a["codigo_pierde"] == "DEST-A"
        assert fila_par_a["fuentes_cambian"] == "SI"
        assert fila_par_a["criterios_cambian"] == "SI"

    def test_par_sin_cambios_reales_se_marca_como_tal(self, sidoe_config, excel_prueba, tmp_path):
        _sembrar(sidoe_config)
        salida = tmp_path / "reporte.csv"
        generar_reporte(excel_prueba, str(salida), db_path=sidoe_config)

        with open(salida, encoding="utf-8-sig") as f:
            filas = list(csv.DictReader(f))

        fila_par_b = next(
            r for r in filas if {r["codigo_gana"], r["codigo_pierde"]} == {"ORIG-B", "DEST-B"}
        )
        # Mismo score_factibilidad_final Y mismo contenido -> "YA AL DÍA",
        # no hay ganador/perdedor real que reportar.
        assert fila_par_b["estado"] == "YA AL DÍA"

    def test_destino_inexistente_se_reporta_sin_romper(self, sidoe_config, excel_prueba, tmp_path):
        _sembrar(sidoe_config)
        salida = tmp_path / "reporte.csv"
        generar_reporte(excel_prueba, str(salida), db_path=sidoe_config)

        with open(salida, encoding="utf-8-sig") as f:
            filas = list(csv.DictReader(f))

        fila_c = next(r for r in filas if "ORIG-C" in (r["codigo_gana"], r["codigo_pierde"]))
        assert "SIN MATCH" in fila_c["estado"]
        assert "NO-EXISTE-999" in fila_c["estado"]

    def test_excel_inexistente_devuelve_none(self, sidoe_config, tmp_path):
        resultado = generar_reporte(str(tmp_path / "no_existe.xlsx"), db_path=sidoe_config)
        assert resultado is None

    def test_empate_con_lado_ods_se_reporta_con_motivo(self, sidoe_config, excel_prueba, tmp_path):
        _sembrar(sidoe_config)
        salida = tmp_path / "reporte.csv"
        generar_reporte(excel_prueba, str(salida), db_path=sidoe_config)

        with open(salida, encoding="utf-8-sig") as f:
            filas = list(csv.DictReader(f))

        fila = next(
            r for r in filas
            if {r["codigo_gana"], r["codigo_pierde"]} == {"CMV-EMPATE", "ODS-EMPATE"}
        )
        assert fila["codigo_gana"] == "ODS-EMPATE"
        assert fila["motivo"] == "empate_desempatado_por_ods"
